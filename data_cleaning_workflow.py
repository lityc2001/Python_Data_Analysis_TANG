"""日常数据清洗与验证工作流（单文件版）。

适用环境：Windows、PyCharm、Python 3.11+、pandas、numpy、openpyxl。
脚本不会修改输入文件；所有结果都写入单独的输出目录。
"""

from __future__ import annotations

# =============================================================================
# 1. 导入依赖
# =============================================================================

import argparse
import copy
import json
import logging
import re
import sys
import tempfile
import time
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# =============================================================================
# 2. 用户配置区（日常使用主要修改这里）
# =============================================================================

# 输入文件路径。可填写绝对路径，也可填写相对于本脚本的路径。
# 留空时必须通过命令行 --input 指定；只运行自检时可以继续留空。
INPUT_PATH = "dirty_data_200000.csv"

# 输出目录。相对路径会以脚本所在目录为基准，而不是以 PyCharm 当前目录为基准。
OUTPUT_DIR = "output"

# Excel 工作表：0 表示第一个工作表，也可以填写工作表名称，例如 "Sheet1"。
SHEET_NAME: int | str = 0

# CSV 编码。None 表示按 utf-8-sig、utf-8、gbk、cp932 顺序自动回退。
# 如果明确知道编码，可以填写 "utf-8-sig"、"gbk" 等。
CSV_ENCODING: str | None = None

# 必须存在的列。缺少时产生 critical 级 schema 问题。
REQUIRED_COLUMNS: list[str] = []

# 允许存在但不是必需的列。配置了必需/可选列后，其他列会被标记为未预期列。
OPTIONAL_COLUMNS: list[str] = []

# 列名重命名映射。列名会先清理空格和换行，再应用这里的映射。
COLUMN_RENAME_MAP: dict[str, str] = {}

# 期望数据类型。未配置的列保持 pandas 读取时的类型，不会被强制转换。
# 支持：string、Int64、float、boolean、category、datetime。
EXPECTED_DTYPES: dict[str, str] = {
    # "material_id": "string",
    # "quantity": "float",
    # "posting_date": "datetime",
}

# 日期列允许的候选格式。程序会按顺序尝试，全部失败才判定为非法日期。
DATE_FORMATS: dict[str, list[str]] = {
    # "posting_date": ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"],
}

# 分类字段允许值。清洗及 VALUE_MAPPING 映射后仍不在列表中的值会报错。
ALLOWED_VALUES: dict[str, list[Any]] = {
    # "status": ["Active", "Inactive"],
}

# 分类值标准化映射。只修改明确列出的值，不猜测业务含义。
VALUE_MAPPING: dict[str, dict[Any, Any]] = {
    # "status": {"active": "Active", "ACTIVE": "Active"},
}

# 数值业务范围。min 或 max 可以只配置其中一个；边界值本身允许出现。
NUMERIC_RANGES: dict[str, dict[str, float | int | None]] = {
    # "quantity": {"min": 0, "max": 100000},
}

# 正则规则使用 fullmatch，即整个单元格都必须符合表达式。
REGEX_RULES: dict[str, str] = {
    # "material_id": r"^[A-Za-z0-9_-]+$",
}

# 业务重复键。例如 ["material_id", "posting_date"]。
DUPLICATE_KEYS: list[str] = []

# 不允许为空的字段。必须先记录缺失问题，之后才会应用缺失值填充策略。
NOT_NULL_COLUMNS: list[str] = []

# 使用 IQR 检测异常值的数值列。
OUTLIER_COLUMNS: list[str] = []

# 缺失值处理策略。支持 none、constant、mean、median、mode、
# forward_fill、backward_fill。未配置时等同于 none。
MISSING_VALUE_STRATEGIES: dict[str, dict[str, Any]] = {
    # "quantity": {"method": "median"},
    # "status": {"method": "constant", "value": "Unknown"},
}

# 可选跨字段规则。operator 支持 <、<=、>、>=、==、!=。
# right 可以是另一个列名；如果与列名不匹配，则按常量比较。
CROSS_FIELD_RULES: list[dict[str, Any]] = [
    # {
    #     "name": "start_before_end",
    #     "left": "start_date",
    #     "operator": "<=",
    #     "right": "end_date",
    #     "message": "开始日期不能晚于结束日期",
    #     "level": "error",
    #     "reject_row": True,
    # }
]

# True：删除完全重复组中 keep 规则之外的记录，并将被删记录放入 rejected_data。
REMOVE_EXACT_DUPLICATES = True

# True：删除业务键重复组中 keep 规则之外的记录；默认仅标记、不删除。
REMOVE_KEY_DUPLICATES = False

# 重复保留规则，只能是 "first" 或 "last"。
EXACT_DUPLICATE_KEEP = "first"
KEY_DUPLICATE_KEEP = "first"

# True：异常值只产生 warning；False：异常值产生 error 并拒绝该行，但仍不改值。
FLAG_OUTLIERS_ONLY = True

# True：除 Excel 报告外，再输出四个 CSV 文件。
EXPORT_CSV_FILES = False

# True：运行内置自检。也可以通过命令行 --self-test 临时开启。
RUN_SELF_TEST = False

# 以下通常无需修改：分类频次展示数量、IQR 系数、自动修改明细采样上限。
TOP_VALUE_COUNT = 10
IQR_MULTIPLIER = 1.5
MAX_CHANGE_DETAIL_ROWS = 100_000
MAX_EXCEL_CELL_LENGTH = 32_000
MAX_EXCEL_DATA_ROWS = 1_048_575  # Excel 总行上限减去一行表头。


# =============================================================================
# 3. 数据模型
# =============================================================================

INTERNAL_ROW_NUMBER = "__row_number__"
ISSUE_COLUMNS = [
    "row_number",
    "rule_name",
    "error_code",
    "field_names",
    "original_value",
    "cleaned_value",
    "message",
    "level",
    "reject_row",
    "category",
]
ISSUE_LEVELS = frozenset({"warning", "error", "critical"})
ISSUE_CATEGORIES = frozenset(
    {"schema", "missing", "type", "format", "duplicate", "business", "outlier"}
)
NA_STRING_TOKENS = frozenset({"", " ", "NA", "N/A", "NULL", "null", "None", "-", "--"})
SUPPORTED_DTYPES = frozenset({"string", "int64", "float", "boolean", "category", "datetime"})


@dataclass(slots=True)
class WorkflowConfig:
    """运行时配置快照，避免核心流程直接修改顶部的用户配置对象。"""

    input_path: Path | None = None
    output_dir: Path = field(default_factory=lambda: Path("output"))
    sheet_name: int | str = 0
    csv_encoding: str | None = None
    required_columns: list[str] = field(default_factory=list)
    optional_columns: list[str] = field(default_factory=list)
    column_rename_map: dict[str, str] = field(default_factory=dict)
    expected_dtypes: dict[str, str] = field(default_factory=dict)
    date_formats: dict[str, list[str]] = field(default_factory=dict)
    allowed_values: dict[str, list[Any]] = field(default_factory=dict)
    value_mapping: dict[str, dict[Any, Any]] = field(default_factory=dict)
    numeric_ranges: dict[str, dict[str, float | int | None]] = field(default_factory=dict)
    regex_rules: dict[str, str] = field(default_factory=dict)
    duplicate_keys: list[str] = field(default_factory=list)
    not_null_columns: list[str] = field(default_factory=list)
    outlier_columns: list[str] = field(default_factory=list)
    missing_value_strategies: dict[str, dict[str, Any]] = field(default_factory=dict)
    cross_field_rules: list[dict[str, Any]] = field(default_factory=list)
    remove_exact_duplicates: bool = True
    remove_key_duplicates: bool = False
    exact_duplicate_keep: str = "first"
    key_duplicate_keep: str = "first"
    flag_outliers_only: bool = True
    export_csv_files: bool = False
    top_value_count: int = 10
    iqr_multiplier: float = 1.5
    max_change_detail_rows: int = 100_000

    def __post_init__(self) -> None:
        """校验不会因数据内容变化的配置错误。"""
        for name, value in (
            ("EXACT_DUPLICATE_KEEP", self.exact_duplicate_keep),
            ("KEY_DUPLICATE_KEEP", self.key_duplicate_keep),
        ):
            if value not in {"first", "last"}:
                raise ValueError(f"{name} 只能是 'first' 或 'last'，当前为：{value!r}")
        if self.iqr_multiplier <= 0:
            raise ValueError("IQR_MULTIPLIER 必须大于 0。")
        if self.top_value_count <= 0:
            raise ValueError("TOP_VALUE_COUNT 必须大于 0。")

    @classmethod
    def from_user_config(cls) -> "WorkflowConfig":
        """从脚本顶部配置创建深拷贝，运行中不会回写用户配置。"""
        script_dir = Path(__file__).resolve().parent
        input_path = _resolve_config_path(INPUT_PATH, script_dir) if INPUT_PATH else None
        output_dir = _resolve_config_path(OUTPUT_DIR, script_dir)
        return cls(
            input_path=input_path,
            output_dir=output_dir,
            sheet_name=SHEET_NAME,
            csv_encoding=CSV_ENCODING,
            required_columns=copy.deepcopy(REQUIRED_COLUMNS),
            optional_columns=copy.deepcopy(OPTIONAL_COLUMNS),
            column_rename_map=copy.deepcopy(COLUMN_RENAME_MAP),
            expected_dtypes=copy.deepcopy(EXPECTED_DTYPES),
            date_formats=copy.deepcopy(DATE_FORMATS),
            allowed_values=copy.deepcopy(ALLOWED_VALUES),
            value_mapping=copy.deepcopy(VALUE_MAPPING),
            numeric_ranges=copy.deepcopy(NUMERIC_RANGES),
            regex_rules=copy.deepcopy(REGEX_RULES),
            duplicate_keys=copy.deepcopy(DUPLICATE_KEYS),
            not_null_columns=copy.deepcopy(NOT_NULL_COLUMNS),
            outlier_columns=copy.deepcopy(OUTLIER_COLUMNS),
            missing_value_strategies=copy.deepcopy(MISSING_VALUE_STRATEGIES),
            cross_field_rules=copy.deepcopy(CROSS_FIELD_RULES),
            remove_exact_duplicates=REMOVE_EXACT_DUPLICATES,
            remove_key_duplicates=REMOVE_KEY_DUPLICATES,
            exact_duplicate_keep=EXACT_DUPLICATE_KEEP,
            key_duplicate_keep=KEY_DUPLICATE_KEEP,
            flag_outliers_only=FLAG_OUTLIERS_ONLY,
            export_csv_files=EXPORT_CSV_FILES,
            top_value_count=TOP_VALUE_COUNT,
            iqr_multiplier=IQR_MULTIPLIER,
            max_change_detail_rows=MAX_CHANGE_DETAIL_ROWS,
        )


@dataclass(frozen=True, slots=True)
class ValidationIssue:
    """统一数据质量问题模型。"""

    row_number: int | None
    rule_name: str
    error_code: str
    field_names: str
    original_value: Any
    cleaned_value: Any
    message: str
    level: str
    reject_row: bool
    category: str

    def unique_key(self) -> tuple[Any, str, str, str]:
        """返回问题去重键。"""
        return (self.row_number, self.rule_name, self.field_names, self.error_code)


@dataclass(slots=True)
class ReadResult:
    """统一文件读取结果。"""

    dataframe: pd.DataFrame
    used_encoding: str | None
    sheet_name: int | str | None


@dataclass(slots=True)
class AssessmentResult:
    """清洗前数据评估表集合。"""

    summary: pd.DataFrame
    column_profile: pd.DataFrame
    numeric_statistics: pd.DataFrame
    value_distribution: pd.DataFrame


@dataclass(slots=True)
class WorkflowResult:
    """一次内存数据工作流的全部结果。"""

    original_data: pd.DataFrame
    cleaned_data: pd.DataFrame
    clean_data: pd.DataFrame
    warning_data: pd.DataFrame
    rejected_data: pd.DataFrame
    issues: pd.DataFrame
    assessment_summary: pd.DataFrame
    column_profile: pd.DataFrame
    numeric_statistics: pd.DataFrame
    value_distribution: pd.DataFrame
    missing_summary: pd.DataFrame
    duplicate_summary: pd.DataFrame
    type_conversion: pd.DataFrame
    outlier_summary: pd.DataFrame
    cleaning_changes: pd.DataFrame


class ExcelWriteError(RuntimeError):
    """Excel 报告写出或二次更新失败。"""


class IssueCollector:
    """集中收集问题，并在添加时按稳定键去重。"""

    def __init__(self) -> None:
        self._issues: list[ValidationIssue] = []
        self._keys: set[tuple[Any, str, str, str]] = set()

    def __len__(self) -> int:
        return len(self._issues)

    def add(
        self,
        *,
        row_number: int | float | np.integer[Any] | None,
        rule_name: str,
        error_code: str,
        field_names: str | Sequence[str],
        original_value: Any,
        cleaned_value: Any,
        message: str,
        level: str,
        reject_row: bool,
        category: str,
    ) -> None:
        """添加一条问题；相同行、规则、字段、代码只保留一次。"""
        if level not in ISSUE_LEVELS:
            raise ValueError(f"无效问题级别：{level}")
        if category not in ISSUE_CATEGORIES:
            raise ValueError(f"无效问题类别：{category}")
        normalized_row = None if row_number is None or pd.isna(row_number) else int(row_number)
        normalized_fields = (
            field_names if isinstance(field_names, str) else ", ".join(map(str, field_names))
        )
        issue = ValidationIssue(
            row_number=normalized_row,
            rule_name=rule_name,
            error_code=error_code,
            field_names=normalized_fields,
            original_value=_display_value(original_value),
            cleaned_value=_display_value(cleaned_value),
            message=message,
            level=level,
            reject_row=bool(reject_row),
            category=category,
        )
        key = issue.unique_key()
        if key not in self._keys:
            self._keys.add(key)
            self._issues.append(issue)

    def add_rows(
        self,
        *,
        row_numbers: Iterable[Any],
        rule_name: str,
        error_code: str,
        field_names: str | Sequence[str],
        original_values: Iterable[Any],
        cleaned_values: Iterable[Any],
        message: str,
        level: str,
        reject_row: bool,
        category: str,
    ) -> None:
        """批量添加同一规则的问题。"""
        for row_number, original_value, cleaned_value in zip(
            row_numbers, original_values, cleaned_values, strict=True
        ):
            self.add(
                row_number=row_number,
                rule_name=rule_name,
                error_code=error_code,
                field_names=field_names,
                original_value=original_value,
                cleaned_value=cleaned_value,
                message=message,
                level=level,
                reject_row=reject_row,
                category=category,
            )

    def to_dataframe(self) -> pd.DataFrame:
        """转换为字段顺序固定的 DataFrame。"""
        if not self._issues:
            return pd.DataFrame(columns=ISSUE_COLUMNS)
        result = pd.DataFrame([asdict(issue) for issue in self._issues], columns=ISSUE_COLUMNS)
        result["row_number"] = pd.array(result["row_number"], dtype="Int64")
        result["reject_row"] = result["reject_row"].astype(bool)
        return result


class ChangeTracker:
    """记录自动修改的汇总与有限明细，避免大数据下无限占用内存。"""

    COLUMNS = [
        "record_type",
        "row_number",
        "field_name",
        "action",
        "original_value",
        "cleaned_value",
        "affected_count",
        "note",
    ]

    def __init__(self, max_details: int) -> None:
        self.max_details = max(0, max_details)
        self._detail_frames: list[pd.DataFrame] = []
        self._detail_count = 0
        self._summary: dict[tuple[str, str], int] = {}

    def record_column_change(self, old_name: str, new_name: str, action: str) -> None:
        """记录列名级修改。"""
        self._summary[(action, new_name)] = self._summary.get((action, new_name), 0) + 1
        if self._detail_count < self.max_details:
            self._detail_frames.append(
                pd.DataFrame(
                    [
                        {
                            "record_type": "detail",
                            "row_number": pd.NA,
                            "field_name": new_name,
                            "action": action,
                            "original_value": old_name,
                            "cleaned_value": new_name,
                            "affected_count": 1,
                            "note": "列名修改",
                        }
                    ]
                )
            )
            self._detail_count += 1

    def record_series_changes(
        self,
        row_numbers: pd.Series,
        field_name: str,
        before: pd.Series,
        after: pd.Series,
        action: str,
    ) -> int:
        """向量化识别变化，并记录汇总和最多 max_details 条明细。"""
        same = (before.eq(after) | (before.isna() & after.isna())).fillna(False)
        changed = ~same
        count = int(changed.sum())
        if count == 0:
            return 0
        self._summary[(action, field_name)] = self._summary.get((action, field_name), 0) + count
        remaining = self.max_details - self._detail_count
        if remaining > 0:
            selected_index = before.index[changed][:remaining]
            details = pd.DataFrame(
                {
                    "record_type": "detail",
                    "row_number": row_numbers.loc[selected_index].to_numpy(),
                    "field_name": field_name,
                    "action": action,
                    "original_value": before.loc[selected_index].to_numpy(),
                    "cleaned_value": after.loc[selected_index].to_numpy(),
                    "affected_count": 1,
                    "note": "",
                }
            )
            self._detail_frames.append(details)
            self._detail_count += len(details)
        return count

    def to_dataframe(self) -> pd.DataFrame:
        """生成先汇总、后明细的自动修改追踪表。"""
        summary_rows = [
            {
                "record_type": "summary",
                "row_number": pd.NA,
                "field_name": field_name,
                "action": action,
                "original_value": "",
                "cleaned_value": "",
                "affected_count": count,
                "note": "自动修改总数",
            }
            for (action, field_name), count in sorted(self._summary.items())
        ]
        frames: list[pd.DataFrame] = [pd.DataFrame(summary_rows, columns=self.COLUMNS)]
        frames.extend(self._detail_frames)
        result = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(columns=self.COLUMNS)
        total_changes = sum(self._summary.values())
        if total_changes > self._detail_count:
            truncation_row = pd.DataFrame(
                [
                    {
                        "record_type": "notice",
                        "row_number": pd.NA,
                        "field_name": "",
                        "action": "DETAIL_TRUNCATED",
                        "original_value": "",
                        "cleaned_value": "",
                        "affected_count": total_changes - self._detail_count,
                        "note": f"为控制内存，仅保留前 {self.max_details} 条修改明细；汇总计数完整。",
                    }
                ]
            )
            result = pd.concat([result, truncation_row], ignore_index=True)
        if result.empty:
            return pd.DataFrame(columns=self.COLUMNS)
        result["row_number"] = pd.array(result["row_number"], dtype="Int64")
        return result[self.COLUMNS]


class MemoryLogHandler(logging.Handler):
    """把 logging 记录保存在内存中，供 Run_Log 工作表写出。"""

    def __init__(self, records: list[dict[str, str]]) -> None:
        super().__init__()
        self.records = records

    def emit(self, record: logging.LogRecord) -> None:
        try:
            message = self.format(record)
            self.records.append(
                {
                    "timestamp": datetime.fromtimestamp(record.created).strftime("%Y-%m-%d %H:%M:%S"),
                    "level": record.levelname,
                    "message": message,
                }
            )
        except (ValueError, TypeError, AttributeError):
            self.handleError(record)


# =============================================================================
# 4. 文件读取
# =============================================================================


def _resolve_config_path(value: str | Path, script_dir: Path | None = None) -> Path:
    """把相对配置路径稳定解析到脚本目录。"""
    path = Path(value).expanduser()
    if path.is_absolute():
        return path.resolve()
    base = script_dir or Path(__file__).resolve().parent
    return (base / path).resolve()


def configure_logging(console: bool = True) -> tuple[logging.Logger, list[dict[str, str]]]:
    """创建同时面向控制台与内存报表的 logger，不使用全局可变业务状态。"""
    records: list[dict[str, str]] = []
    logger = logging.getLogger(f"data_cleaning_workflow.{time.time_ns()}")
    logger.setLevel(logging.INFO)
    logger.propagate = False
    formatter = logging.Formatter("%(message)s")
    if console:
        stream_handler = logging.StreamHandler(sys.stdout)
        stream_handler.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        logger.addHandler(stream_handler)
    memory_handler = MemoryLogHandler(records)
    memory_handler.setFormatter(formatter)
    logger.addHandler(memory_handler)
    return logger, records


def read_input_file(path: Path, config: WorkflowConfig) -> ReadResult:
    """统一读取 CSV、XLSX 或 XLS，添加原文件行号但不修改输入文件。"""
    resolved_path = path.expanduser().resolve()
    if not resolved_path.exists():
        raise FileNotFoundError(f"输入文件不存在：{resolved_path}")
    if not resolved_path.is_file():
        raise ValueError(f"输入路径不是文件：{resolved_path}")
    if resolved_path.stat().st_size == 0:
        raise ValueError(f"输入文件为空（0 字节）：{resolved_path}")

    suffix = resolved_path.suffix.lower()
    used_encoding: str | None = None
    sheet_name: int | str | None = None

    if suffix == ".csv":
        fallback_encodings = ["utf-8-sig", "utf-8", "gbk", "cp932"]
        encodings = ([config.csv_encoding] if config.csv_encoding else []) + fallback_encodings
        encodings = list(dict.fromkeys(encoding for encoding in encodings if encoding))
        last_unicode_error: UnicodeDecodeError | None = None
        dataframe: pd.DataFrame | None = None
        for encoding in encodings:
            try:
                dataframe = pd.read_csv(resolved_path, encoding=encoding)
                used_encoding = encoding
                break
            except UnicodeDecodeError as exc:
                last_unicode_error = exc
            except pd.errors.EmptyDataError as exc:
                raise ValueError(f"CSV 文件没有表头或数据：{resolved_path}") from exc
            except pd.errors.ParserError as exc:
                raise ValueError(f"CSV 结构无法解析：{resolved_path}；{exc}") from exc
        if dataframe is None:
            if last_unicode_error is not None:
                raise last_unicode_error
            raise ValueError(f"CSV 文件读取失败：{resolved_path}")
    elif suffix in {".xlsx", ".xls"}:
        sheet_name = config.sheet_name
        try:
            dataframe = pd.read_excel(resolved_path, sheet_name=sheet_name)
        except ImportError as exc:
            if suffix == ".xls":
                raise ImportError(
                    "读取旧版 .xls 需要 pandas 环境提供 xlrd 引擎；建议另存为 .xlsx，"
                    "或在允许的环境中安装 xlrd。"
                ) from exc
            raise
        except ValueError as exc:
            raise ValueError(f"Excel 工作表读取失败（sheet={sheet_name!r}）：{exc}") from exc
    else:
        raise ValueError(f"不支持的文件类型：{suffix or '无扩展名'}；仅支持 .csv、.xlsx、.xls。")

    if not isinstance(dataframe, pd.DataFrame):
        raise ValueError("读取结果不是单个 DataFrame，请检查 SHEET_NAME 配置。")
    if dataframe.shape[1] == 0:
        raise ValueError(f"输入文件没有任何列：{resolved_path}")
    if dataframe.shape[0] == 0:
        raise ValueError(f"输入文件只有表头，没有数据行：{resolved_path}")
    if INTERNAL_ROW_NUMBER in dataframe.columns:
        raise ValueError(f"输入文件包含保留列名 {INTERNAL_ROW_NUMBER!r}，请先重命名该业务列。")

    # 表头占第 1 行，因此 CSV/Excel 第一条数据的原始行号统一为 2。
    dataframe = dataframe.copy(deep=True)
    dataframe[INTERNAL_ROW_NUMBER] = np.arange(2, len(dataframe) + 2, dtype=np.int64)
    return ReadResult(dataframe=dataframe, used_encoding=used_encoding, sheet_name=sheet_name)


# =============================================================================
# 5. 数据评估
# =============================================================================


def _data_columns(dataframe: pd.DataFrame) -> list[Any]:
    """返回排除内部行号的业务列名。"""
    return [column for column in dataframe.columns if column != INTERNAL_ROW_NUMBER]


def _display_value(value: Any) -> Any:
    """把复杂值转换为报告可读形式，同时保留常见标量类型。"""
    if value is None or value is pd.NA:
        return None
    if isinstance(value, (dict, list, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    if isinstance(value, np.generic):
        return value.item()
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        return str(value)
    return value


def parse_datetime_series(series: pd.Series, formats: Sequence[str] | None = None) -> pd.Series:
    """按多个候选格式解析日期；不修改传入 Series。"""
    parsed = pd.Series(pd.NaT, index=series.index, dtype="datetime64[ns]")
    non_missing = series.notna()
    if formats:
        for date_format in formats:
            remaining = non_missing & parsed.isna()
            if not remaining.any():
                break
            candidate = pd.to_datetime(series.loc[remaining], format=date_format, errors="coerce")
            parsed.loc[remaining] = candidate
        return parsed
    try:
        candidate = pd.to_datetime(series.loc[non_missing], errors="coerce", format="mixed")
    except (TypeError, ValueError):
        candidate = pd.to_datetime(series.loc[non_missing], errors="coerce")
    parsed.loc[non_missing] = candidate
    return parsed


def evaluate_data(
    dataframe: pd.DataFrame,
    config: WorkflowConfig,
    *,
    file_name: str,
    sheet_name: int | str | None,
) -> AssessmentResult:
    """在不删除、不改写数据的前提下生成初始评估。"""
    business_positions = [
        position
        for position, column in enumerate(dataframe.columns)
        if column != INTERNAL_ROW_NUMBER
    ]
    source_columns = [str(dataframe.columns[position]) for position in business_positions]
    columns = _make_unique_columns(source_columns)
    business_data = dataframe.iloc[:, business_positions].copy(deep=True)
    business_data.columns = columns
    initial_duplicate_names = sorted(
        {column for column in source_columns if source_columns.count(column) > 1}
    )
    expected_set = set(config.required_columns) | set(config.optional_columns)
    unexpected_columns = [column for column in columns if expected_set and column not in expected_set]
    missing_required = [column for column in config.required_columns if column not in columns]
    fully_empty = [column for column in columns if business_data[column].isna().all()]
    suspected_constant = [
        column
        for column in columns
        if business_data[column].notna().any() and business_data[column].nunique(dropna=True) <= 1
    ]
    exact_duplicate_count = int(business_data.duplicated(keep="first").sum()) if columns else 0
    existing_keys = [key for key in config.duplicate_keys if key in business_data.columns]
    key_duplicate_count = (
        int(business_data.duplicated(subset=existing_keys, keep=False).sum())
        if existing_keys and len(existing_keys) == len(config.duplicate_keys)
        else 0
    )

    summary_items: list[tuple[str, Any]] = [
        ("文件名", file_name),
        ("工作表名", sheet_name if sheet_name is not None else "CSV"),
        ("总行数", len(dataframe)),
        ("总列数", len(columns)),
        ("列名", ", ".join(source_columns)),
        ("初始重复列名", ", ".join(initial_duplicate_names)),
        ("完全重复行数量（后续重复）", exact_duplicate_count),
        ("关键字段重复数量（组内全部记录）", key_duplicate_count),
        ("完全为空的列", ", ".join(map(str, fully_empty))),
        ("疑似常量列", ", ".join(map(str, suspected_constant))),
        ("未预期出现的列", ", ".join(map(str, unexpected_columns))),
        ("缺失的必需列", ", ".join(missing_required)),
    ]
    summary = pd.DataFrame(summary_items, columns=["指标", "值"])

    profile_rows: list[dict[str, Any]] = []
    date_columns = {
        column
        for column, dtype_name in config.expected_dtypes.items()
        if dtype_name.lower() == "datetime"
    } | set(config.date_formats)
    for column in columns:
        series = business_data[column]
        missing_count = int(series.isna().sum())
        date_min: Any = None
        date_max: Any = None
        if column in date_columns or pd.api.types.is_datetime64_any_dtype(series):
            parsed_dates = (
                series if pd.api.types.is_datetime64_any_dtype(series) else parse_datetime_series(series, config.date_formats.get(column))
            )
            if parsed_dates.notna().any():
                date_min = parsed_dates.min()
                date_max = parsed_dates.max()
        profile_rows.append(
            {
                "column_name": str(column),
                "dtype": str(series.dtype),
                "non_null_count": int(series.notna().sum()),
                "missing_count": missing_count,
                "missing_rate": round(missing_count / len(series), 6) if len(series) else 0.0,
                "unique_count": int(series.nunique(dropna=True)),
                "is_fully_empty": bool(series.isna().all()),
                "is_suspected_constant": bool(
                    series.notna().any() and series.nunique(dropna=True) <= 1
                ),
                "is_unexpected": bool(expected_set and column not in expected_set),
                "date_min": date_min,
                "date_max": date_max,
            }
        )
    column_profile = pd.DataFrame(profile_rows)

    numeric_columns = list(business_data.select_dtypes(include=[np.number]).columns)
    if numeric_columns:
        numeric_statistics = business_data[numeric_columns].describe().transpose().reset_index()
        numeric_statistics = numeric_statistics.rename(columns={"index": "column_name"})
    else:
        numeric_statistics = pd.DataFrame(
            columns=["column_name", "count", "mean", "std", "min", "25%", "50%", "75%", "max"]
        )

    distribution_rows: list[dict[str, Any]] = []
    categorical_columns = [
        column
        for column in columns
        if not pd.api.types.is_numeric_dtype(business_data[column])
        and not pd.api.types.is_datetime64_any_dtype(business_data[column])
    ]
    for column in categorical_columns:
        counts = business_data[column].value_counts(dropna=False).head(config.top_value_count)
        for rank, (value, count) in enumerate(counts.items(), start=1):
            distribution_rows.append(
                {
                    "column_name": str(column),
                    "rank": rank,
                    "value": "<缺失>" if pd.isna(value) else _display_value(value),
                    "count": int(count),
                    "percentage": round(int(count) / len(dataframe), 6) if len(dataframe) else 0.0,
                }
            )
    value_distribution = pd.DataFrame(
        distribution_rows,
        columns=["column_name", "rank", "value", "count", "percentage"],
    )
    return AssessmentResult(
        summary=summary,
        column_profile=column_profile,
        numeric_statistics=numeric_statistics,
        value_distribution=value_distribution,
    )


# =============================================================================
# 6. 通用数据清洗
# =============================================================================


def normalize_column_name(column_name: Any) -> str:
    """清理列名前后空格、换行与连续空白。"""
    normalized = re.sub(r"[\r\n]+", " ", str(column_name))
    return re.sub(r"\s+", " ", normalized).strip()


def _make_unique_columns(columns: Sequence[str]) -> list[str]:
    """为重复列名添加可预测后缀，保证后续流程可安全引用。"""
    seen: dict[str, int] = {}
    unique_columns: list[str] = []
    for column in columns:
        occurrence = seen.get(column, 0) + 1
        seen[column] = occurrence
        unique_columns.append(column if occurrence == 1 else f"{column}__duplicate_{occurrence}")
    return unique_columns


def _series_string_mask(series: pd.Series) -> pd.Series:
    """识别 object 列中真正的字符串值，避免把数字对象误转成文本。"""
    return series.map(lambda value: isinstance(value, str), na_action="ignore").fillna(False).astype(bool)


def clean_general_data(
    dataframe: pd.DataFrame,
    config: WorkflowConfig,
    issues: IssueCollector,
    changes: ChangeTracker,
) -> tuple[pd.DataFrame, dict[str, Any]]:
    """执行列名、字符串、缺失标记和明确值映射清洗。"""
    cleaned = dataframe.copy(deep=True)
    original_columns = list(cleaned.columns)
    business_original_columns = [column for column in original_columns if column != INTERNAL_ROW_NUMBER]

    normalized_columns: list[str] = []
    lineage_before_unique: list[Any] = []
    for original_column in business_original_columns:
        normalized = normalize_column_name(original_column)
        renamed = config.column_rename_map.get(normalized, normalized)
        normalized_columns.append(normalize_column_name(renamed))
        lineage_before_unique.append(original_column)
        if str(original_column) != normalized:
            changes.record_column_change(str(original_column), normalized, "COLUMN_NAME_CLEAN")
        if normalized != normalize_column_name(renamed):
            changes.record_column_change(normalized, normalize_column_name(renamed), "COLUMN_RENAME")
        elif normalized != normalize_column_name(original_column):
            changes.record_column_change(normalize_column_name(original_column), normalized, "COLUMN_RENAME")

    duplicate_names = sorted(
        {column for column in normalized_columns if normalized_columns.count(column) > 1}
    )
    for duplicate_name in duplicate_names:
        issues.add(
            row_number=None,
            rule_name="duplicate_column_names_after_rename",
            error_code="DUPLICATE_COLUMN_NAME",
            field_names=duplicate_name,
            original_value=duplicate_name,
            cleaned_value=None,
            message=f"列名清理或重命名后出现重复列名：{duplicate_name}。已加后缀以便继续生成报告。",
            level="critical",
            reject_row=False,
            category="schema",
        )

    unique_business_columns = _make_unique_columns(normalized_columns)
    for before, after in zip(normalized_columns, unique_business_columns, strict=True):
        if before != after:
            changes.record_column_change(before, after, "DUPLICATE_COLUMN_DISAMBIGUATION")
    cleaned.columns = unique_business_columns + [INTERNAL_ROW_NUMBER]
    lineage = dict(zip(unique_business_columns, lineage_before_unique, strict=True))

    for column in unique_business_columns:
        series = cleaned[column]
        if not (
            pd.api.types.is_object_dtype(series)
            or isinstance(series.dtype, pd.StringDtype)
            or isinstance(series.dtype, pd.CategoricalDtype)
        ):
            continue
        if isinstance(series.dtype, pd.CategoricalDtype):
            series = series.astype(object)
            cleaned[column] = series
        string_mask = _series_string_mask(series)
        whitespace_cleaned = series.copy()
        if string_mask.any():
            normalized_text = (
                series.loc[string_mask]
                .astype("string")
                .str.strip()
                .str.replace(r"\s+", " ", regex=True)
            )
            whitespace_cleaned.loc[string_mask] = normalized_text.astype(object)
        changes.record_series_changes(
            cleaned[INTERNAL_ROW_NUMBER],
            column,
            series,
            whitespace_cleaned,
            "STRING_WHITESPACE_CLEAN",
        )
        cleaned[column] = whitespace_cleaned

        missing_marked = cleaned[column].copy()
        missing_token_mask = _series_string_mask(missing_marked) & missing_marked.isin(NA_STRING_TOKENS)
        missing_marked.loc[missing_token_mask] = pd.NA
        changes.record_series_changes(
            cleaned[INTERNAL_ROW_NUMBER],
            column,
            cleaned[column],
            missing_marked,
            "MISSING_TOKEN_TO_NULL",
        )
        cleaned[column] = missing_marked

    for column, mapping in config.value_mapping.items():
        if column not in cleaned.columns:
            continue
        before = cleaned[column].copy()
        try:
            mapping_mask = before.isin(list(mapping.keys()))
        except TypeError:
            mapping_mask = before.map(
                lambda value: value in mapping if isinstance(value, (str, int, float, bool, tuple)) else False
            )
        after = before.copy()
        if mapping_mask.any():
            after.loc[mapping_mask] = before.loc[mapping_mask].map(mapping)
        changes.record_series_changes(
            cleaned[INTERNAL_ROW_NUMBER], column, before, after, "VALUE_MAPPING"
        )
        cleaned[column] = after
    return cleaned, lineage


def _lookup_original_values(
    original_data: pd.DataFrame,
    cleaned_column: str,
    row_numbers: Sequence[Any] | pd.Series | np.ndarray[Any, Any],
    lineage: Mapping[str, Any],
    fallback: Sequence[Any] | pd.Series | np.ndarray[Any, Any],
) -> list[Any]:
    """按内部行号回查输入值；无法唯一回查时使用调用方提供的值。"""
    original_column = lineage.get(cleaned_column)
    if original_column not in original_data.columns or INTERNAL_ROW_NUMBER not in original_data.columns:
        return list(fallback)
    original_series = original_data.set_index(INTERNAL_ROW_NUMBER)[original_column]
    if isinstance(original_series, pd.DataFrame):
        original_series = original_series.iloc[:, 0]
    try:
        return original_series.reindex(pd.Index(row_numbers)).tolist()
    except (KeyError, ValueError, TypeError):
        return list(fallback)


# =============================================================================
# 7. 数据类型转换
# =============================================================================


def _convert_boolean_series(series: pd.Series) -> pd.Series:
    """把常见且含义明确的布尔表达转换为 pandas 可空 boolean。"""
    result = pd.Series(pd.NA, index=series.index, dtype="boolean")
    true_tokens = frozenset({"true", "t", "yes", "y", "1", "是", "真"})
    false_tokens = frozenset({"false", "f", "no", "n", "0", "否", "假"})
    non_missing = series.notna()
    boolean_values = non_missing & series.map(lambda value: isinstance(value, (bool, np.bool_)))
    if boolean_values.any():
        result.loc[boolean_values] = series.loc[boolean_values].astype(bool)
    numeric_values = non_missing & series.map(
        lambda value: isinstance(value, (int, float, np.integer, np.floating))
        and not isinstance(value, (bool, np.bool_))
    )
    if numeric_values.any():
        numeric = pd.to_numeric(series.loc[numeric_values], errors="coerce")
        result.loc[numeric.index[numeric.eq(1)]] = True
        result.loc[numeric.index[numeric.eq(0)]] = False
    string_values = _series_string_mask(series)
    if string_values.any():
        normalized = series.loc[string_values].astype("string").str.strip().str.lower()
        result.loc[normalized.index[normalized.isin(true_tokens)]] = True
        result.loc[normalized.index[normalized.isin(false_tokens)]] = False
    return result


def convert_data_types(
    dataframe: pd.DataFrame,
    original_data: pd.DataFrame,
    lineage: Mapping[str, Any],
    config: WorkflowConfig,
    issues: IssueCollector,
    changes: ChangeTracker,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """按 EXPECTED_DTYPES 转换类型，失败单元格转为空并记录问题。"""
    converted_data = dataframe.copy(deep=True)
    statistics: list[dict[str, Any]] = []

    for column, configured_dtype in config.expected_dtypes.items():
        normalized_dtype = configured_dtype.lower()
        if column not in converted_data.columns:
            statistics.append(
                {
                    "column_name": column,
                    "configured_dtype": configured_dtype,
                    "before_dtype": "<列不存在>",
                    "after_dtype": "<列不存在>",
                    "non_null_before": 0,
                    "non_null_after": 0,
                    "failure_count": 0,
                    "status": "skipped_missing_column",
                }
            )
            continue
        if normalized_dtype not in SUPPORTED_DTYPES:
            issues.add(
                row_number=None,
                rule_name="supported_expected_dtype",
                error_code="UNSUPPORTED_DTYPE",
                field_names=column,
                original_value=configured_dtype,
                cleaned_value=None,
                message=f"列 {column} 配置了不支持的数据类型：{configured_dtype}。",
                level="critical",
                reject_row=False,
                category="schema",
            )
            statistics.append(
                {
                    "column_name": column,
                    "configured_dtype": configured_dtype,
                    "before_dtype": str(converted_data[column].dtype),
                    "after_dtype": str(converted_data[column].dtype),
                    "non_null_before": int(converted_data[column].notna().sum()),
                    "non_null_after": int(converted_data[column].notna().sum()),
                    "failure_count": 0,
                    "status": "unsupported_dtype",
                }
            )
            continue

        before = converted_data[column].copy()
        before_dtype = str(before.dtype)
        non_missing_before = before.notna()
        try:
            if normalized_dtype == "string":
                converted = before.astype("string")
            elif normalized_dtype == "float":
                numeric = pd.to_numeric(before, errors="coerce")
                non_finite = numeric.notna() & ~np.isfinite(numeric.astype(float))
                numeric.loc[non_finite] = np.nan
                converted = numeric.astype(float)
            elif normalized_dtype == "int64":
                numeric = pd.to_numeric(before, errors="coerce")
                non_finite = numeric.notna() & ~np.isfinite(numeric.astype(float))
                fractional = numeric.notna() & ~np.isclose(numeric.astype(float) % 1, 0)
                numeric.loc[non_finite | fractional] = np.nan
                converted = numeric.astype("Int64")
            elif normalized_dtype == "boolean":
                converted = _convert_boolean_series(before)
            elif normalized_dtype == "category":
                converted = before.astype("category")
            else:  # datetime
                converted = parse_datetime_series(before, config.date_formats.get(column))
        except (TypeError, ValueError, OverflowError) as exc:
            issues.add(
                row_number=None,
                rule_name="column_type_conversion",
                error_code="COLUMN_TYPE_CONVERSION_ERROR",
                field_names=column,
                original_value=before_dtype,
                cleaned_value=None,
                message=f"列 {column} 类型转换阶段发生错误：{exc}",
                level="critical",
                reject_row=False,
                category="type",
            )
            statistics.append(
                {
                    "column_name": column,
                    "configured_dtype": configured_dtype,
                    "before_dtype": before_dtype,
                    "after_dtype": before_dtype,
                    "non_null_before": int(non_missing_before.sum()),
                    "non_null_after": int(non_missing_before.sum()),
                    "failure_count": 0,
                    "status": f"conversion_error: {exc}",
                }
            )
            continue

        invalid_mask = non_missing_before & converted.isna()
        invalid_rows = converted_data.loc[invalid_mask, INTERNAL_ROW_NUMBER]
        invalid_before = before.loc[invalid_mask]
        original_values = _lookup_original_values(
            original_data,
            column,
            invalid_rows.to_numpy(),
            lineage,
            invalid_before.to_numpy(),
        )
        issues.add_rows(
            row_numbers=invalid_rows.to_numpy(),
            rule_name="expected_dtype_conversion",
            error_code="TYPE_CONVERSION_FAILED",
            field_names=column,
            original_values=original_values,
            cleaned_values=[None] * len(invalid_rows),
            message=f"字段 {column} 无法转换为配置类型 {configured_dtype}。",
            level="error",
            reject_row=True,
            category="type",
        )
        if normalized_dtype == "datetime" and invalid_mask.any():
            issues.add_rows(
                row_numbers=invalid_rows.to_numpy(),
                rule_name="allowed_date_format",
                error_code="DATE_FORMAT_INVALID",
                field_names=column,
                original_values=original_values,
                cleaned_values=[None] * len(invalid_rows),
                message=f"字段 {column} 不符合允许的日期格式。",
                level="error",
                reject_row=True,
                category="format",
            )

        converted_data[column] = converted
        changes.record_series_changes(
            converted_data[INTERNAL_ROW_NUMBER], column, before, converted, "TYPE_CONVERSION"
        )
        statistics.append(
            {
                "column_name": column,
                "configured_dtype": configured_dtype,
                "before_dtype": before_dtype,
                "after_dtype": str(converted.dtype),
                "non_null_before": int(non_missing_before.sum()),
                "non_null_after": int(converted.notna().sum()),
                "failure_count": int(invalid_mask.sum()),
                "status": "completed",
            }
        )
    type_conversion = pd.DataFrame(
        statistics,
        columns=[
            "column_name",
            "configured_dtype",
            "before_dtype",
            "after_dtype",
            "non_null_before",
            "non_null_after",
            "failure_count",
            "status",
        ],
    )
    return converted_data, type_conversion


# =============================================================================
# 8. 数据验证
# =============================================================================


def _original_values_for_mask(
    original_data: pd.DataFrame,
    dataframe: pd.DataFrame,
    column: str,
    mask: pd.Series,
    lineage: Mapping[str, Any],
) -> list[Any]:
    """为布尔掩码对应的问题回查输入文件原值。"""
    rows = dataframe.loc[mask, INTERNAL_ROW_NUMBER]
    fallback = dataframe.loc[mask, column] if column in dataframe.columns else pd.Series([None] * len(rows))
    return _lookup_original_values(
        original_data, column, rows.to_numpy(), lineage, fallback.to_numpy()
    )


def validate_schema(
    dataframe: pd.DataFrame,
    config: WorkflowConfig,
    issues: IssueCollector,
) -> None:
    """验证必需列、重复列名和完全空列。"""
    business_columns = _data_columns(dataframe)
    missing_required = [column for column in config.required_columns if column not in business_columns]
    for column in missing_required:
        issues.add(
            row_number=None,
            rule_name="required_column_exists",
            error_code="MISSING_REQUIRED_COLUMN",
            field_names=column,
            original_value=None,
            cleaned_value=None,
            message=f"缺少必需列：{column}。",
            level="critical",
            reject_row=False,
            category="schema",
        )

    duplicate_names = [
        str(column) for column in dataframe.columns[dataframe.columns.duplicated(keep=False)].unique()
    ]
    for column in duplicate_names:
        issues.add(
            row_number=None,
            rule_name="unique_column_names",
            error_code="DUPLICATE_COLUMN_NAME",
            field_names=column,
            original_value=column,
            cleaned_value=None,
            message=f"存在重复列名：{column}。",
            level="critical",
            reject_row=False,
            category="schema",
        )

    for column in business_columns:
        if dataframe[column].isna().all():
            required_empty = column in config.required_columns
            issues.add(
                row_number=None,
                rule_name="column_not_fully_empty",
                error_code="FULLY_EMPTY_COLUMN",
                field_names=column,
                original_value=None,
                cleaned_value=None,
                message=f"列 {column} 完全为空。",
                level="error" if required_empty else "warning",
                reject_row=False,
                category="schema",
            )


def validate_missing_values(
    dataframe: pd.DataFrame,
    original_data: pd.DataFrame,
    lineage: Mapping[str, Any],
    config: WorkflowConfig,
    issues: IssueCollector,
) -> None:
    """在任何填充发生前记录必填字段缺失问题。"""
    for column in config.not_null_columns:
        if column not in dataframe.columns:
            continue
        missing_mask = dataframe[column].isna()
        if not missing_mask.any():
            continue
        rows = dataframe.loc[missing_mask, INTERNAL_ROW_NUMBER]
        originals = _original_values_for_mask(
            original_data, dataframe, column, missing_mask, lineage
        )
        issues.add_rows(
            row_numbers=rows.to_numpy(),
            rule_name="required_value_not_null",
            error_code="REQUIRED_VALUE_MISSING",
            field_names=column,
            original_values=originals,
            cleaned_values=[None] * len(rows),
            message=f"必填字段 {column} 为空。",
            level="error",
            reject_row=True,
            category="missing",
        )


def validate_regex_rules(
    dataframe: pd.DataFrame,
    original_data: pd.DataFrame,
    lineage: Mapping[str, Any],
    config: WorkflowConfig,
    issues: IssueCollector,
) -> None:
    """按 fullmatch 方式验证正则格式。"""
    for column, pattern in config.regex_rules.items():
        if column not in dataframe.columns:
            continue
        try:
            compiled = re.compile(pattern)
        except re.error as exc:
            issues.add(
                row_number=None,
                rule_name="valid_regex_configuration",
                error_code="INVALID_REGEX_RULE",
                field_names=column,
                original_value=pattern,
                cleaned_value=None,
                message=f"字段 {column} 的正则表达式无效：{exc}",
                level="critical",
                reject_row=False,
                category="schema",
            )
            continue
        series = dataframe[column]
        matches = series.astype("string").str.fullmatch(compiled, na=True).fillna(True)
        invalid_mask = series.notna() & ~matches.astype(bool)
        if not invalid_mask.any():
            continue
        rows = dataframe.loc[invalid_mask, INTERNAL_ROW_NUMBER]
        originals = _original_values_for_mask(
            original_data, dataframe, column, invalid_mask, lineage
        )
        issues.add_rows(
            row_numbers=rows.to_numpy(),
            rule_name="regex_fullmatch",
            error_code="REGEX_MISMATCH",
            field_names=column,
            original_values=originals,
            cleaned_values=dataframe.loc[invalid_mask, column].to_numpy(),
            message=f"字段 {column} 不符合正则格式：{pattern}",
            level="error",
            reject_row=True,
            category="format",
        )


def validate_allowed_values(
    dataframe: pd.DataFrame,
    original_data: pd.DataFrame,
    lineage: Mapping[str, Any],
    config: WorkflowConfig,
    issues: IssueCollector,
) -> None:
    """验证分类值是否位于显式允许集合。"""
    for column, allowed in config.allowed_values.items():
        if column not in dataframe.columns:
            continue
        series = dataframe[column]
        invalid_mask = series.notna() & ~series.isin(allowed)
        if not invalid_mask.any():
            continue
        rows = dataframe.loc[invalid_mask, INTERNAL_ROW_NUMBER]
        originals = _original_values_for_mask(
            original_data, dataframe, column, invalid_mask, lineage
        )
        issues.add_rows(
            row_numbers=rows.to_numpy(),
            rule_name="allowed_value_set",
            error_code="ALLOWED_VALUE_VIOLATION",
            field_names=column,
            original_values=originals,
            cleaned_values=series.loc[invalid_mask].to_numpy(),
            message=f"字段 {column} 出现不允许的分类值；允许值为：{allowed}。",
            level="error",
            reject_row=True,
            category="business",
        )


def validate_numeric_ranges(
    dataframe: pd.DataFrame,
    original_data: pd.DataFrame,
    lineage: Mapping[str, Any],
    config: WorkflowConfig,
    issues: IssueCollector,
) -> None:
    """验证数值是否位于配置的闭区间。"""
    for column, range_rule in config.numeric_ranges.items():
        if column not in dataframe.columns:
            continue
        numeric = pd.to_numeric(dataframe[column], errors="coerce")
        minimum = range_rule.get("min")
        maximum = range_rule.get("max")
        invalid_mask = pd.Series(False, index=dataframe.index)
        if minimum is not None:
            invalid_mask |= numeric.notna() & numeric.lt(minimum)
        if maximum is not None:
            invalid_mask |= numeric.notna() & numeric.gt(maximum)
        if not invalid_mask.any():
            continue
        rows = dataframe.loc[invalid_mask, INTERNAL_ROW_NUMBER]
        originals = _original_values_for_mask(
            original_data, dataframe, column, invalid_mask, lineage
        )
        issues.add_rows(
            row_numbers=rows.to_numpy(),
            rule_name="numeric_range",
            error_code="NUMERIC_RANGE_VIOLATION",
            field_names=column,
            original_values=originals,
            cleaned_values=dataframe.loc[invalid_mask, column].to_numpy(),
            message=f"字段 {column} 超出允许范围 [{minimum}, {maximum}]。",
            level="error",
            reject_row=True,
            category="business",
        )


def validate_cross_field_rules(
    dataframe: pd.DataFrame,
    config: WorkflowConfig,
    issues: IssueCollector,
) -> None:
    """执行配置驱动的简单跨字段比较规则，作为业务规则扩展点。"""
    operators = {
        "<": lambda left, right: left < right,
        "<=": lambda left, right: left <= right,
        ">": lambda left, right: left > right,
        ">=": lambda left, right: left >= right,
        "==": lambda left, right: left == right,
        "!=": lambda left, right: left != right,
    }
    for index, rule in enumerate(config.cross_field_rules, start=1):
        name = str(rule.get("name", f"cross_field_rule_{index}"))
        left_name = rule.get("left")
        operator_name = rule.get("operator")
        right_spec = rule.get("right")
        if left_name not in dataframe.columns or operator_name not in operators:
            issues.add(
                row_number=None,
                rule_name=name,
                error_code="INVALID_CROSS_FIELD_RULE",
                field_names=str(left_name),
                original_value=rule,
                cleaned_value=None,
                message=f"跨字段规则 {name} 配置无效或左字段不存在。",
                level="critical",
                reject_row=False,
                category="schema",
            )
            continue
        left = dataframe[left_name]
        if isinstance(right_spec, str) and right_spec in dataframe.columns:
            right: pd.Series | Any = dataframe[right_spec]
            valid_inputs = left.notna() & right.notna()
            field_names = [str(left_name), right_spec]
        else:
            right = right_spec
            valid_inputs = left.notna() & (right_spec is not None)
            field_names = [str(left_name)]
        try:
            comparison = operators[str(operator_name)](left, right)
            invalid_mask = valid_inputs & ~pd.Series(comparison, index=dataframe.index).fillna(False)
        except (TypeError, ValueError) as exc:
            issues.add(
                row_number=None,
                rule_name=name,
                error_code="CROSS_FIELD_RULE_EVALUATION_ERROR",
                field_names=field_names,
                original_value=rule,
                cleaned_value=None,
                message=f"跨字段规则 {name} 无法执行：{exc}",
                level="critical",
                reject_row=False,
                category="business",
            )
            continue
        level = str(rule.get("level", "error"))
        if level not in ISSUE_LEVELS:
            level = "error"
        reject_row = bool(rule.get("reject_row", level != "warning"))
        message = str(rule.get("message", f"跨字段规则 {name} 不满足。"))
        for row_index in dataframe.index[invalid_mask]:
            cleaned_values = {field: dataframe.at[row_index, field] for field in field_names}
            issues.add(
                row_number=dataframe.at[row_index, INTERNAL_ROW_NUMBER],
                rule_name=name,
                error_code="CROSS_FIELD_RULE_VIOLATION",
                field_names=field_names,
                original_value=cleaned_values,
                cleaned_value=cleaned_values,
                message=message,
                level=level,
                reject_row=reject_row,
                category="business",
            )


def validate_data(
    dataframe: pd.DataFrame,
    original_data: pd.DataFrame,
    lineage: Mapping[str, Any],
    config: WorkflowConfig,
    issues: IssueCollector,
) -> None:
    """执行 schema、missing、format 与 business 验证。"""
    validate_schema(dataframe, config, issues)
    validate_missing_values(dataframe, original_data, lineage, config, issues)
    validate_regex_rules(dataframe, original_data, lineage, config, issues)
    validate_allowed_values(dataframe, original_data, lineage, config, issues)
    validate_numeric_ranges(dataframe, original_data, lineage, config, issues)
    validate_cross_field_rules(dataframe, config, issues)


# =============================================================================
# 9. 重复数据处理
# =============================================================================


def _duplicate_group_hash(dataframe: pd.DataFrame, columns: list[str]) -> pd.Series:
    """为重复组生成稳定哈希；极少数不可哈希对象回退到字符串表示。"""
    try:
        return pd.util.hash_pandas_object(dataframe[columns], index=False)
    except TypeError:
        return pd.util.hash_pandas_object(dataframe[columns].astype("string"), index=False)


def process_duplicates(
    dataframe: pd.DataFrame,
    config: WorkflowConfig,
    issues: IssueCollector,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """先记录完整重复与业务键重复组，再按开关删除后续/前序记录。"""
    working = dataframe.copy(deep=True)
    business_columns = _data_columns(working)
    summary_rows: list[dict[str, Any]] = []
    rows_to_remove: set[int] = set()

    if business_columns:
        exact_all = working.duplicated(subset=business_columns, keep=False)
        exact_remove = (
            working.duplicated(subset=business_columns, keep=config.exact_duplicate_keep)
            if config.remove_exact_duplicates
            else pd.Series(False, index=working.index)
        )
        exact_removed_rows = set(
            working.loc[exact_remove, INTERNAL_ROW_NUMBER].astype(int).tolist()
        )
        rows_to_remove.update(exact_removed_rows)
        if exact_all.any():
            candidates = working.loc[exact_all, [INTERNAL_ROW_NUMBER] + business_columns].copy()
            candidates["__group_hash__"] = _duplicate_group_hash(candidates, business_columns)
            for group_index, (_, group) in enumerate(
                candidates.groupby("__group_hash__", sort=False, dropna=False), start=1
            ):
                if len(group) < 2:
                    continue
                group_id = f"EXACT-{group_index:06d}"
                row_numbers = group[INTERNAL_ROW_NUMBER].astype(int).tolist()
                removed = [row for row in row_numbers if row in exact_removed_rows]
                values = group.iloc[0][business_columns].to_dict()
                summary_rows.append(
                    {
                        "group_type": "exact",
                        "group_id": group_id,
                        "field_names": "<全部业务列>",
                        "key_values": _display_value(values),
                        "row_numbers": ", ".join(map(str, row_numbers)),
                        "count": len(row_numbers),
                        "remove_enabled": config.remove_exact_duplicates,
                        "keep": config.exact_duplicate_keep,
                        "removed_row_numbers": ", ".join(map(str, removed)),
                    }
                )
                for row_number in row_numbers:
                    rejected = row_number in exact_removed_rows
                    issues.add(
                        row_number=row_number,
                        rule_name="exact_duplicate_rows",
                        error_code="EXACT_DUPLICATE",
                        field_names=business_columns,
                        original_value=values,
                        cleaned_value=values,
                        message=f"完全重复组 {group_id}，原始行号：{row_numbers}。",
                        level="error" if rejected else "warning",
                        reject_row=rejected,
                        category="duplicate",
                    )

    if config.duplicate_keys:
        missing_keys = [key for key in config.duplicate_keys if key not in working.columns]
        if missing_keys:
            issues.add(
                row_number=None,
                rule_name="duplicate_key_columns_exist",
                error_code="DUPLICATE_KEY_COLUMN_MISSING",
                field_names=missing_keys,
                original_value=missing_keys,
                cleaned_value=None,
                message=f"业务重复键字段不存在：{missing_keys}。",
                level="critical",
                reject_row=False,
                category="schema",
            )
        else:
            key_all = working.duplicated(subset=config.duplicate_keys, keep=False)
            key_remove = (
                working.duplicated(subset=config.duplicate_keys, keep=config.key_duplicate_keep)
                if config.remove_key_duplicates
                else pd.Series(False, index=working.index)
            )
            key_removed_rows = set(
                working.loc[key_remove, INTERNAL_ROW_NUMBER].astype(int).tolist()
            )
            rows_to_remove.update(key_removed_rows)
            if key_all.any():
                candidates = working.loc[key_all]
                grouping: Any = (
                    candidates.groupby(config.duplicate_keys[0], sort=False, dropna=False)
                    if len(config.duplicate_keys) == 1
                    else candidates.groupby(config.duplicate_keys, sort=False, dropna=False)
                )
                for group_index, (key_values, group) in enumerate(grouping, start=1):
                    if len(group) < 2:
                        continue
                    group_id = f"KEY-{group_index:06d}"
                    row_numbers = group[INTERNAL_ROW_NUMBER].astype(int).tolist()
                    removed = [row for row in row_numbers if row in key_removed_rows]
                    key_dict = group.iloc[0][config.duplicate_keys].to_dict()
                    summary_rows.append(
                        {
                            "group_type": "key",
                            "group_id": group_id,
                            "field_names": ", ".join(config.duplicate_keys),
                            "key_values": _display_value(key_dict),
                            "row_numbers": ", ".join(map(str, row_numbers)),
                            "count": len(row_numbers),
                            "remove_enabled": config.remove_key_duplicates,
                            "keep": config.key_duplicate_keep,
                            "removed_row_numbers": ", ".join(map(str, removed)),
                        }
                    )
                    for row_number in row_numbers:
                        rejected = row_number in key_removed_rows
                        issues.add(
                            row_number=row_number,
                            rule_name="business_key_duplicates",
                            error_code="KEY_DUPLICATE",
                            field_names=config.duplicate_keys,
                            original_value=key_dict,
                            cleaned_value=key_dict,
                            message=f"业务键重复组 {group_id}，原始行号：{row_numbers}。",
                            level="error" if rejected else "warning",
                            reject_row=rejected,
                            category="duplicate",
                        )

    if rows_to_remove:
        working = working.loc[~working[INTERNAL_ROW_NUMBER].isin(rows_to_remove)].copy()
    duplicate_summary = pd.DataFrame(
        summary_rows,
        columns=[
            "group_type",
            "group_id",
            "field_names",
            "key_values",
            "row_numbers",
            "count",
            "remove_enabled",
            "keep",
            "removed_row_numbers",
        ],
    )
    return working, duplicate_summary


# =============================================================================
# 10. 异常值检测
# =============================================================================


def detect_outliers(
    dataframe: pd.DataFrame,
    original_data: pd.DataFrame,
    lineage: Mapping[str, Any],
    config: WorkflowConfig,
    issues: IssueCollector,
) -> pd.DataFrame:
    """使用 IQR 检测异常值；不删除、不替换、不截尾。"""
    summary_rows: list[dict[str, Any]] = []
    for column in config.outlier_columns:
        if column not in dataframe.columns:
            summary_rows.append(
                {
                    "column_name": column,
                    "valid_count": 0,
                    "q1": None,
                    "q3": None,
                    "iqr": None,
                    "lower_bound": None,
                    "upper_bound": None,
                    "outlier_count": 0,
                    "status": "skipped_missing_column",
                }
            )
            continue
        numeric = pd.to_numeric(dataframe[column], errors="coerce")
        valid = numeric.dropna()
        if len(valid) < 4:
            summary_rows.append(
                {
                    "column_name": column,
                    "valid_count": len(valid),
                    "q1": None,
                    "q3": None,
                    "iqr": None,
                    "lower_bound": None,
                    "upper_bound": None,
                    "outlier_count": 0,
                    "status": "skipped_too_few_values",
                }
            )
            continue
        q1 = float(valid.quantile(0.25))
        q3 = float(valid.quantile(0.75))
        iqr = q3 - q1
        if np.isclose(iqr, 0.0):
            summary_rows.append(
                {
                    "column_name": column,
                    "valid_count": len(valid),
                    "q1": q1,
                    "q3": q3,
                    "iqr": iqr,
                    "lower_bound": q1,
                    "upper_bound": q3,
                    "outlier_count": 0,
                    "status": "skipped_zero_iqr",
                }
            )
            continue
        lower = q1 - config.iqr_multiplier * iqr
        upper = q3 + config.iqr_multiplier * iqr
        outlier_mask = numeric.notna() & (numeric.lt(lower) | numeric.gt(upper))
        outlier_count = int(outlier_mask.sum())
        summary_rows.append(
            {
                "column_name": column,
                "valid_count": len(valid),
                "q1": q1,
                "q3": q3,
                "iqr": iqr,
                "lower_bound": lower,
                "upper_bound": upper,
                "outlier_count": outlier_count,
                "status": "flagged" if config.flag_outliers_only else "rejected_without_value_change",
            }
        )
        if not outlier_mask.any():
            continue
        rows = dataframe.loc[outlier_mask, INTERNAL_ROW_NUMBER]
        originals = _original_values_for_mask(
            original_data, dataframe, column, outlier_mask, lineage
        )
        issues.add_rows(
            row_numbers=rows.to_numpy(),
            rule_name="iqr_outlier_detection",
            error_code="IQR_OUTLIER",
            field_names=column,
            original_values=originals,
            cleaned_values=dataframe.loc[outlier_mask, column].to_numpy(),
            message=f"字段 {column} 超出 IQR 边界 [{lower:.6g}, {upper:.6g}]，值未被修改。",
            level="warning" if config.flag_outliers_only else "error",
            reject_row=not config.flag_outliers_only,
            category="outlier",
        )
    return pd.DataFrame(
        summary_rows,
        columns=[
            "column_name",
            "valid_count",
            "q1",
            "q3",
            "iqr",
            "lower_bound",
            "upper_bound",
            "outlier_count",
            "status",
        ],
    )


# =============================================================================
# 11. 缺失值处理与数据分组
# =============================================================================


def _fill_missing_series(
    series: pd.Series,
    strategy: Mapping[str, Any],
) -> tuple[pd.Series, str]:
    """应用一个字段的缺失策略，返回新 Series 与状态说明。"""
    method = str(strategy.get("method", "none")).lower()
    if method == "none":
        return series.copy(), "not_filled"
    if method == "constant":
        if "value" not in strategy:
            return series.copy(), "skipped_constant_without_value"
        fill_value = strategy.get("value")
    elif method in {"mean", "median"}:
        numeric = pd.to_numeric(series, errors="coerce")
        valid = numeric.dropna()
        if valid.empty:
            return series.copy(), "skipped_no_numeric_value"
        fill_value = float(valid.mean() if method == "mean" else valid.median())
        if str(series.dtype) == "Int64" and not float(fill_value).is_integer():
            return series.copy(), "skipped_non_integer_fill_for_Int64"
        if str(series.dtype) == "Int64":
            fill_value = int(fill_value)
    elif method == "mode":
        modes = series.mode(dropna=True)
        if modes.empty:
            return series.copy(), "skipped_no_mode"
        fill_value = modes.iloc[0]
    elif method == "forward_fill":
        return series.ffill(), "completed"
    elif method == "backward_fill":
        return series.bfill(), "completed"
    else:
        return series.copy(), f"skipped_unknown_method:{method}"

    try:
        if isinstance(series.dtype, pd.CategoricalDtype) and fill_value not in series.cat.categories:
            series = series.cat.add_categories([fill_value])
        return series.fillna(fill_value), "completed"
    except (TypeError, ValueError) as exc:
        return series.copy(), f"skipped_incompatible_fill:{exc}"


def apply_missing_value_strategies(
    dataframe: pd.DataFrame,
    config: WorkflowConfig,
    changes: ChangeTracker,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """在缺失问题记录完成后执行显式填充策略，不删除任何行。"""
    filled_data = dataframe.copy(deep=True)
    summary_rows: list[dict[str, Any]] = []
    configured_missing_columns = set(config.missing_value_strategies)
    columns = _data_columns(filled_data)
    for column in columns:
        before = filled_data[column].copy()
        before_count = int(before.isna().sum())
        strategy = config.missing_value_strategies.get(column, {"method": "none"})
        after, status = _fill_missing_series(before, strategy)
        fill_mask = before.isna() & after.notna()
        filled_count = int(fill_mask.sum())
        if filled_count:
            changes.record_series_changes(
                filled_data[INTERNAL_ROW_NUMBER], column, before, after, "MISSING_VALUE_FILL"
            )
            filled_data[column] = after
        summary_rows.append(
            {
                "column_name": column,
                "missing_before": before_count,
                "missing_rate_before": round(before_count / len(before), 6) if len(before) else 0.0,
                "strategy": str(strategy.get("method", "none")),
                "configured_value": _display_value(strategy.get("value")),
                "filled_count": filled_count,
                "missing_after": int(filled_data[column].isna().sum()),
                "status": status,
            }
        )
    for column in sorted(configured_missing_columns - set(columns)):
        strategy = config.missing_value_strategies[column]
        summary_rows.append(
            {
                "column_name": column,
                "missing_before": 0,
                "missing_rate_before": 0.0,
                "strategy": str(strategy.get("method", "none")),
                "configured_value": _display_value(strategy.get("value")),
                "filled_count": 0,
                "missing_after": 0,
                "status": "skipped_missing_column",
            }
        )
    return filled_data, pd.DataFrame(
        summary_rows,
        columns=[
            "column_name",
            "missing_before",
            "missing_rate_before",
            "strategy",
            "configured_value",
            "filled_count",
            "missing_after",
            "status",
        ],
    )


def _rename_internal_row_number_for_output(dataframe: pd.DataFrame) -> pd.DataFrame:
    """把内部行号改为 row_number，并避免覆盖同名业务列。"""
    output = dataframe.copy(deep=True)
    if "row_number" in output.columns and INTERNAL_ROW_NUMBER in output.columns:
        replacement = "source_row_number"
        suffix = 2
        while replacement in output.columns:
            replacement = f"source_row_number_{suffix}"
            suffix += 1
        output = output.rename(columns={"row_number": replacement})
    return output.rename(columns={INTERNAL_ROW_NUMBER: "row_number"})


def split_data_by_issues(
    cleaned_data: pd.DataFrame,
    original_data: pd.DataFrame,
    issues_dataframe: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """按 reject_row 和 warning 划分互斥/子集关系明确的三类数据。"""
    if issues_dataframe.empty:
        reject_rows: set[int] = set()
        warning_rows: set[int] = set()
    else:
        row_issues = issues_dataframe.loc[issues_dataframe["row_number"].notna()].copy()
        reject_rows = set(
            row_issues.loc[row_issues["reject_row"].astype(bool), "row_number"].astype(int).tolist()
        )
        warning_rows = set(
            row_issues.loc[row_issues["level"].eq("warning"), "row_number"].astype(int).tolist()
        )

    clean_internal = cleaned_data.loc[
        ~cleaned_data[INTERNAL_ROW_NUMBER].isin(reject_rows)
    ].copy()
    warning_internal = clean_internal.loc[
        clean_internal[INTERNAL_ROW_NUMBER].isin(warning_rows)
    ].copy()
    rejected_internal = original_data.loc[
        original_data[INTERNAL_ROW_NUMBER].isin(reject_rows)
    ].copy()

    clean_data = _rename_internal_row_number_for_output(clean_internal)
    warning_data = _rename_internal_row_number_for_output(warning_internal)
    rejected_data = _rename_internal_row_number_for_output(rejected_internal)
    _assert_partition_relationships(clean_data, warning_data, rejected_data)
    return clean_data, warning_data, rejected_data


def _assert_partition_relationships(
    clean_data: pd.DataFrame,
    warning_data: pd.DataFrame,
    rejected_data: pd.DataFrame,
) -> None:
    """运行时验证 clean/rejected 互斥且 warning 是 clean 子集。"""
    clean_rows = set(clean_data.get("row_number", pd.Series(dtype="int64")).dropna().astype(int))
    warning_rows = set(warning_data.get("row_number", pd.Series(dtype="int64")).dropna().astype(int))
    rejected_rows = set(rejected_data.get("row_number", pd.Series(dtype="int64")).dropna().astype(int))
    if clean_rows & rejected_rows:
        raise RuntimeError("数据分组失败：clean_data 与 rejected_data 存在重叠行号。")
    if not warning_rows.issubset(clean_rows):
        raise RuntimeError("数据分组失败：warning_data 不是 clean_data 的子集。")


def _ensure_internal_row_numbers(dataframe: pd.DataFrame) -> pd.DataFrame:
    """为内存 DataFrame 添加或校验内部原始行号，并把内部列移到最后。"""
    prepared = dataframe.copy(deep=True)
    internal_count = list(prepared.columns).count(INTERNAL_ROW_NUMBER)
    if internal_count > 1:
        raise ValueError(f"内部列 {INTERNAL_ROW_NUMBER} 只能存在一次。")
    if internal_count == 0:
        prepared[INTERNAL_ROW_NUMBER] = np.arange(2, len(prepared) + 2, dtype=np.int64)
    else:
        row_numbers = pd.to_numeric(prepared[INTERNAL_ROW_NUMBER], errors="coerce")
        if row_numbers.isna().any() or row_numbers.duplicated().any():
            raise ValueError(f"内部列 {INTERNAL_ROW_NUMBER} 必须为非空且唯一的数值行号。")
        prepared[INTERNAL_ROW_NUMBER] = row_numbers.astype(np.int64)
    business_positions = [
        position
        for position, column in enumerate(prepared.columns)
        if column != INTERNAL_ROW_NUMBER
    ]
    internal_positions = [
        position
        for position, column in enumerate(prepared.columns)
        if column == INTERNAL_ROW_NUMBER
    ]
    if len(internal_positions) != 1:
        raise ValueError(f"内部列 {INTERNAL_ROW_NUMBER} 只能存在一次。")
    return prepared.iloc[:, business_positions + internal_positions].copy(deep=True)


def run_dataframe_workflow(
    dataframe: pd.DataFrame,
    config: WorkflowConfig,
    *,
    source_name: str = "<memory>",
    sheet_name: int | str | None = None,
    logger: logging.Logger | None = None,
) -> WorkflowResult:
    """对一个内存 DataFrame 执行完整工作流，供 main() 和自检共同使用。"""
    active_logger = logger or configure_logging(console=False)[0]
    if dataframe.shape[1] == 0:
        raise ValueError("输入 DataFrame 没有任何列。")
    if dataframe.shape[0] == 0:
        raise ValueError("输入 DataFrame 没有任何数据行。")
    prepared = _ensure_internal_row_numbers(dataframe)
    original_data = prepared.copy(deep=True)

    active_logger.info("阶段 1/8 开始：初始数据评估，行数=%s", len(prepared))
    assessment = evaluate_data(
        original_data,
        config,
        file_name=source_name,
        sheet_name=sheet_name,
    )
    active_logger.info("阶段 1/8 结束：初始数据评估完成，列数=%s", len(_data_columns(prepared)))

    issues = IssueCollector()
    changes = ChangeTracker(config.max_change_detail_rows)

    active_logger.info("阶段 2/8 开始：通用数据清洗")
    cleaned, lineage = clean_general_data(prepared, config, issues, changes)
    active_logger.info("阶段 2/8 结束：通用清洗完成，行数=%s，累计问题=%s", len(cleaned), len(issues))

    active_logger.info("阶段 3/8 开始：数据类型转换")
    converted, type_conversion = convert_data_types(
        cleaned, original_data, lineage, config, issues, changes
    )
    active_logger.info("阶段 3/8 结束：类型转换完成，累计问题=%s", len(issues))

    active_logger.info("阶段 4/8 开始：schema、缺失、格式和业务规则验证")
    validate_data(converted, original_data, lineage, config, issues)
    active_logger.info("阶段 4/8 结束：规则验证完成，累计问题=%s", len(issues))

    active_logger.info("阶段 5/8 开始：重复数据检查与配置化处理")
    deduplicated, duplicate_summary = process_duplicates(converted, config, issues)
    active_logger.info(
        "阶段 5/8 结束：重复处理后行数=%s，累计问题=%s", len(deduplicated), len(issues)
    )

    active_logger.info("阶段 6/8 开始：IQR 异常值检测")
    outlier_summary = detect_outliers(
        deduplicated, original_data, lineage, config, issues
    )
    active_logger.info("阶段 6/8 结束：异常值检测完成，累计问题=%s", len(issues))

    active_logger.info("阶段 7/8 开始：缺失值策略处理")
    filled, missing_summary = apply_missing_value_strategies(deduplicated, config, changes)
    active_logger.info("阶段 7/8 结束：缺失值处理完成，行数=%s", len(filled))

    active_logger.info("阶段 8/8 开始：问题汇总与结果分组")
    issues_dataframe = issues.to_dataframe()
    clean_data, warning_data, rejected_data = split_data_by_issues(
        filled, original_data, issues_dataframe
    )
    cleaning_changes = changes.to_dataframe()
    active_logger.info(
        "阶段 8/8 结束：clean=%s，warning=%s，rejected=%s，问题=%s",
        len(clean_data),
        len(warning_data),
        len(rejected_data),
        len(issues_dataframe),
    )
    return WorkflowResult(
        original_data=original_data,
        cleaned_data=filled,
        clean_data=clean_data,
        warning_data=warning_data,
        rejected_data=rejected_data,
        issues=issues_dataframe,
        assessment_summary=assessment.summary,
        column_profile=assessment.column_profile,
        numeric_statistics=assessment.numeric_statistics,
        value_distribution=assessment.value_distribution,
        missing_summary=missing_summary,
        duplicate_summary=duplicate_summary,
        type_conversion=type_conversion,
        outlier_summary=outlier_summary,
        cleaning_changes=cleaning_changes,
    )


# =============================================================================
# 12. Excel 报告生成
# =============================================================================


def calculate_data_quality_score(issues: pd.DataFrame, original_row_count: int) -> float:
    """计算 0~100 的透明质量分数。

    公式：100 - 100 × 加权问题数 / (原始行数 × 10)，最低为 0。
    warning/error/critical 权重分别为 1/3/10。分母中的 10 表示“每行累计到
    一个 critical 等价问题时扣满”，因此规模不同的数据集仍可直观比较。
    该分数只用于日常趋势观察，不替代业务上的验收结论。
    """
    if issues.empty:
        return 100.0
    weights = {"warning": 1.0, "error": 3.0, "critical": 10.0}
    weighted_issues = float(issues["level"].map(weights).fillna(0).sum())
    denominator = max(int(original_row_count), 1) * 10.0
    return round(max(0.0, 100.0 - 100.0 * weighted_issues / denominator), 2)


def build_summary_dataframe(
    result: WorkflowResult,
    *,
    input_path: str | Path,
    output_path: str | Path,
    started_at: datetime,
    elapsed_seconds: float,
) -> pd.DataFrame:
    """生成 Summary 工作表的指标和值。"""
    issues = result.issues

    def count_level(level: str) -> int:
        return int(issues["level"].eq(level).sum()) if not issues.empty else 0

    def count_category(category: str) -> int:
        return int(issues["category"].eq(category).sum()) if not issues.empty else 0

    metrics: list[tuple[str, Any]] = [
        ("输入文件", str(input_path)),
        ("输出文件", str(output_path)),
        ("运行时间", started_at.strftime("%Y-%m-%d %H:%M:%S")),
        ("总运行时长（秒）", round(elapsed_seconds, 3)),
        ("原始行数", len(result.original_data)),
        ("清洗后行数", len(result.cleaned_data)),
        ("clean 行数", len(result.clean_data)),
        ("warning 行数", len(result.warning_data)),
        ("rejected 行数", len(result.rejected_data)),
        ("问题总数", len(issues)),
        ("warning 数", count_level("warning")),
        ("error 数", count_level("error")),
        ("critical 数", count_level("critical")),
        ("缺失问题数", count_category("missing")),
        ("类型问题数", count_category("type")),
        ("重复问题数", count_category("duplicate")),
        ("格式问题数", count_category("format")),
        ("业务规则问题数", count_category("business")),
        ("异常值问题数", count_category("outlier")),
        ("schema 问题数", count_category("schema")),
        ("数据质量分数", calculate_data_quality_score(issues, len(result.original_data))),
        (
            "质量分数公式",
            "max(0, 100 - 100 × (warning×1 + error×3 + critical×10) / (原始行数×10))",
        ),
    ]
    return pd.DataFrame(metrics, columns=["指标", "值"])


def _safe_excel_value(value: Any) -> Any:
    """清理 Excel 不接受的控制字符，并截断超长单元格。"""
    value = _display_value(value)
    if value is None:
        return None
    if isinstance(value, str):
        value = re.sub(r"[\x00-\x08\x0B\x0C\x0E-\x1F]", "", value)
        if len(value) > MAX_EXCEL_CELL_LENGTH:
            value = value[: MAX_EXCEL_CELL_LENGTH - 8] + "…[已截断]"
    return value


def _prepare_excel_dataframe(
    dataframe: pd.DataFrame,
    sheet_name: str,
    logger: logging.Logger,
) -> pd.DataFrame:
    """复制并转换为可由 openpyxl 安全写出的 DataFrame。"""
    if len(dataframe) > MAX_EXCEL_DATA_ROWS:
        logger.warning(
            "工作表 %s 有 %s 行，超过 Excel 上限；报告中仅写前 %s 行。",
            sheet_name,
            len(dataframe),
            MAX_EXCEL_DATA_ROWS,
        )
    prepared = dataframe.head(MAX_EXCEL_DATA_ROWS).copy(deep=True)
    if prepared.shape[1] == 0:
        prepared = pd.DataFrame({"说明": pd.Series(dtype="string")})
    prepared.columns = _make_unique_columns([normalize_column_name(column) for column in prepared.columns])
    prepared = prepared.replace([np.inf, -np.inf], np.nan)
    for column in prepared.columns:
        series = prepared[column]
        if (
            pd.api.types.is_object_dtype(series)
            or isinstance(series.dtype, pd.StringDtype)
            or isinstance(series.dtype, pd.CategoricalDtype)
        ):
            prepared[column] = series.astype(object).map(_safe_excel_value)
        elif isinstance(series.dtype, pd.DatetimeTZDtype):
            prepared[column] = series.dt.tz_localize(None)
    return prepared


def _safe_sheet_name(name: str, used_names: set[str]) -> str:
    """生成符合 Excel 31 字符限制且不重复的工作表名。"""
    base = re.sub(r"[\\/*?:\[\]]", "_", name).strip("'")[:31] or "Sheet"
    candidate = base
    suffix = 2
    while candidate in used_names:
        marker = f"_{suffix}"
        candidate = f"{base[: 31 - len(marker)]}{marker}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def _format_worksheet(worksheet: Any, dataframe: pd.DataFrame) -> None:
    """冻结首行、筛选、表头加粗、日期格式和合理列宽。"""
    worksheet.freeze_panes = "A2"
    if worksheet.max_column > 0 and worksheet.max_row > 0:
        worksheet.auto_filter.ref = worksheet.dimensions
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    sample_end = min(worksheet.max_row, 201)
    for column_index, column_name in enumerate(dataframe.columns, start=1):
        lengths = [len(str(column_name))]
        for row_index in range(2, sample_end + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            lengths.append(len(str(value)) if value is not None else 0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max(lengths) + 2, 10), 50
        )
        if pd.api.types.is_datetime64_any_dtype(dataframe[column_name]):
            for row_index in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_index, column=column_index).number_format = "yyyy-mm-dd hh:mm:ss"


def write_excel_report(
    report_path: Path,
    result: WorkflowResult,
    summary: pd.DataFrame,
    run_log: pd.DataFrame,
    logger: logging.Logger,
) -> None:
    """写出多工作表 xlsx 报告并应用统一格式。"""
    report_frames: list[tuple[str, pd.DataFrame]] = [
        ("Summary", summary),
        ("Clean_Data", result.clean_data),
        ("Warning_Data", result.warning_data),
        ("Rejected_Data", result.rejected_data),
        ("Validation_Issues", result.issues),
        ("Missing_Summary", result.missing_summary),
        ("Duplicate_Summary", result.duplicate_summary),
        ("Type_Conversion", result.type_conversion),
        ("Value_Distribution", result.value_distribution),
        ("Outlier_Summary", result.outlier_summary),
        ("Column_Profile", result.column_profile),
        ("Run_Log", run_log),
        # 下面三个是额外审计表，不影响“至少包含”的十二张规定工作表。
        ("Initial_Assessment", result.assessment_summary),
        ("Numeric_Statistics", result.numeric_statistics),
        ("Cleaning_Changes", result.cleaning_changes),
    ]
    report_path.parent.mkdir(parents=True, exist_ok=True)
    used_names: set[str] = set()
    try:
        with pd.ExcelWriter(
            report_path,
            engine="openpyxl",
            datetime_format="yyyy-mm-dd hh:mm:ss",
            date_format="yyyy-mm-dd",
        ) as writer:
            for requested_name, dataframe in report_frames:
                sheet_name = _safe_sheet_name(requested_name, used_names)
                prepared = _prepare_excel_dataframe(dataframe, sheet_name, logger)
                prepared.to_excel(writer, sheet_name=sheet_name, index=False)
                _format_worksheet(writer.book[sheet_name], prepared)
    except PermissionError:
        raise
    except (OSError, ValueError, TypeError, ImportError) as exc:
        raise ExcelWriteError(f"Excel 报告写出失败：{report_path}；{exc}") from exc


def update_run_log_sheet(
    report_path: Path,
    log_records: list[dict[str, str]],
    logger: logging.Logger,
) -> None:
    """报告初次保存后更新 Run_Log，使其包含写出成功和总耗时日志。"""
    try:
        workbook = load_workbook(report_path)
        if "Run_Log" in workbook.sheetnames:
            old_sheet = workbook["Run_Log"]
            position = workbook.index(old_sheet)
            workbook.remove(old_sheet)
        else:
            position = len(workbook.sheetnames)
        worksheet = workbook.create_sheet("Run_Log", position)
        log_dataframe = _prepare_excel_dataframe(pd.DataFrame(log_records), "Run_Log", logger)
        for row in dataframe_to_rows(log_dataframe, index=False, header=True):
            worksheet.append(row)
        _format_worksheet(worksheet, log_dataframe)
        workbook.save(report_path)
        workbook.close()
    except PermissionError:
        raise
    except (OSError, ValueError, TypeError, KeyError) as exc:
        raise ExcelWriteError(f"Run_Log 工作表更新失败：{report_path}；{exc}") from exc


def export_csv_outputs(
    output_dir: Path,
    result: WorkflowResult,
) -> list[Path]:
    """按开关额外写出四个 UTF-8-SIG CSV 文件。"""
    output_dir.mkdir(parents=True, exist_ok=True)
    frames = {
        "clean_data.csv": result.clean_data,
        "warning_data.csv": result.warning_data,
        "rejected_data.csv": result.rejected_data,
        "validation_issues.csv": result.issues,
    }
    paths: list[Path] = []
    for file_name, dataframe in frames.items():
        path = output_dir / file_name
        dataframe.to_csv(path, index=False, encoding="utf-8-sig")
        paths.append(path)
    return paths


# =============================================================================
# 13. 日志与异常处理、14. 内置自检
# =============================================================================


def _self_test_dataframe() -> pd.DataFrame:
    """创建覆盖正常、缺失、格式、重复、范围和异常值的内存数据。"""
    records: list[dict[str, Any]] = [
        {"id": "OK1", "amount": 10, "posting_date": "2026-01-01", "status": "active", "code": "A_1"},
        {"id": "OK2", "amount": 11, "posting_date": "2026/01/02", "status": "ACTIVE", "code": "A_2"},
        {"id": "OK3", "amount": 12, "posting_date": "20260103", "status": "Inactive", "code": "A_3"},
        {"id": "OK4", "amount": 13, "posting_date": "2026-01-04", "status": "Active", "code": "A_4"},
        {"id": "OK5", "amount": 14, "posting_date": "2026-01-05", "status": "Active", "code": "A_5"},
        {"id": "OUT", "amount": 1000, "posting_date": "2026-01-06", "status": "Active", "code": "OUT"},
        {"id": "BADNUM", "amount": "abc", "posting_date": "2026-01-07", "status": "Active", "code": "BADNUM"},
        {"id": "BADDATE", "amount": 15, "posting_date": "2026-99-99", "status": "Active", "code": "BADDATE"},
        {"id": "MISS", "amount": 16, "posting_date": "2026-01-09", "status": None, "code": "MISS"},
        {"id": "BADCAT", "amount": 17, "posting_date": "2026-01-10", "status": "Pending", "code": "BADCAT"},
        {"id": "BADREG", "amount": 18, "posting_date": "2026-01-11", "status": "Active", "code": "#bad"},
        {"id": "RANGE", "amount": 200, "posting_date": "2026-01-12", "status": "Active", "code": "RANGE"},
        {"id": "KEY", "amount": 19, "posting_date": "2026-01-13", "status": "Active", "code": "KEY1"},
        {"id": "KEY", "amount": 20, "posting_date": "2026-01-14", "status": "Inactive", "code": "KEY2"},
    ]
    records.append(records[1].copy())  # 完全重复，同时也构成业务键重复。
    dataframe = pd.DataFrame(records)
    whole_numbers: list[Any] = [1] * len(dataframe)
    whole_numbers[6] = 1.5  # Int64 非整数转换失败。
    dataframe["whole_number"] = pd.Series(whole_numbers, dtype=object)
    dataframe["flag"] = "yes"
    dataframe.loc[9, "flag"] = "maybe"  # 非法布尔值。
    dataframe["segment"] = "A"
    dataframe["all_empty"] = None
    return dataframe


def _self_test_config(output_dir: Path) -> WorkflowConfig:
    """返回不依赖用户顶部配置的确定性自检配置。"""
    return WorkflowConfig(
        output_dir=output_dir,
        required_columns=["id", "amount", "posting_date", "status", "code", "missing_required"],
        optional_columns=["whole_number", "flag", "segment", "all_empty"],
        expected_dtypes={
            "id": "string",
            "amount": "float",
            "posting_date": "datetime",
            "status": "string",
            "code": "string",
            "whole_number": "Int64",
            "flag": "boolean",
            "segment": "category",
        },
        date_formats={"posting_date": ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"]},
        allowed_values={"status": ["Active", "Inactive"]},
        value_mapping={"status": {"active": "Active", "ACTIVE": "Active"}},
        numeric_ranges={"amount": {"min": 0, "max": 500}},
        regex_rules={"code": r"^[A-Za-z0-9_-]+$"},
        duplicate_keys=["id"],
        not_null_columns=["id", "amount", "posting_date", "status", "code"],
        outlier_columns=["amount"],
        missing_value_strategies={"status": {"method": "constant", "value": "Unknown"}},
        remove_exact_duplicates=True,
        remove_key_duplicates=False,
        exact_duplicate_keep="first",
        key_duplicate_keep="first",
        flag_outliers_only=True,
        export_csv_files=False,
        max_change_detail_rows=10_000,
    )


def run_self_test() -> None:
    """执行单文件自检，并用临时目录验证 CSV、Excel 与报告输出。"""
    logger, log_records = configure_logging(console=False)
    with tempfile.TemporaryDirectory(prefix="data_cleaning_self_test_") as temporary:
        temporary_path = Path(temporary)
        config = _self_test_config(temporary_path)
        source = _self_test_dataframe()
        result = run_dataframe_workflow(
            source,
            config,
            source_name="self_test_memory",
            sheet_name=None,
            logger=logger,
        )

        clean_rows = set(result.clean_data["row_number"].astype(int))
        warning_rows = set(result.warning_data["row_number"].astype(int))
        rejected_rows = set(result.rejected_data["row_number"].astype(int))
        assert clean_rows.isdisjoint(rejected_rows), "clean_data 与 rejected_data 必须互斥"
        assert warning_rows.issubset(clean_rows), "warning_data 必须是 clean_data 子集"
        assert not result.issues.duplicated(
            subset=["row_number", "rule_name", "field_names", "error_code"]
        ).any(), "ValidationIssue 存在重复记录"
        assert result.original_data[INTERNAL_ROW_NUMBER].tolist() == list(
            range(2, len(source) + 2)
        ), "row_number 必须从 2 开始且连续"
        assert 2 in clean_rows and 2 not in rejected_rows, "正常数据不应被错误拒绝"

        expected_error_codes = {
            "MISSING_REQUIRED_COLUMN",
            "REQUIRED_VALUE_MISSING",
            "TYPE_CONVERSION_FAILED",
            "DATE_FORMAT_INVALID",
            "ALLOWED_VALUE_VIOLATION",
            "EXACT_DUPLICATE",
            "KEY_DUPLICATE",
            "NUMERIC_RANGE_VIOLATION",
            "REGEX_MISMATCH",
            "IQR_OUTLIER",
        }
        actual_error_codes = set(result.issues["error_code"])
        missing_codes = expected_error_codes - actual_error_codes
        assert not missing_codes, f"自检未产生预期问题代码：{sorted(missing_codes)}"
        assert (
            result.cleaning_changes["action"].eq("VALUE_MAPPING").any()
        ), "分类值不一致映射应当被 Cleaning_Changes 追踪"
        assert str(result.cleaned_data["whole_number"].dtype) == "Int64"
        assert str(result.cleaned_data["flag"].dtype) == "boolean"
        assert str(result.cleaned_data["segment"].dtype) == "category"
        assert result.issues["error_code"].eq("FULLY_EMPTY_COLUMN").any()
        bad_number_issues = result.issues.loc[
            result.issues["error_code"].eq("TYPE_CONVERSION_FAILED")
            & result.issues["field_names"].eq("amount")
        ]
        assert "abc" in set(bad_number_issues["original_value"]), "类型问题必须保留输入原值"
        assert result.clean_data.loc[
            result.clean_data["row_number"].eq(2), "status"
        ].iat[0] == "Active", "clean_data 应使用映射后的值"
        last_source_row = len(source) + 1
        assert result.rejected_data.loc[
            result.rejected_data["row_number"].eq(last_source_row), "status"
        ].iat[0] == "ACTIVE", "rejected_data 应默认保留输入原值"

        duplicate_name_result = run_dataframe_workflow(
            pd.DataFrame({"left": [1, 2], "right": [3, 4]}),
            WorkflowConfig(column_rename_map={"left": "same", "right": "same"}),
            source_name="duplicate_column_self_test",
            logger=logger,
        )
        assert duplicate_name_result.issues["error_code"].eq("DUPLICATE_COLUMN_NAME").any()
        raw_duplicate_result = run_dataframe_workflow(
            pd.DataFrame([[1, 2], [3, 4]], columns=["raw", "raw"]),
            WorkflowConfig(),
            source_name="raw_duplicate_column_self_test",
            logger=logger,
        )
        assert raw_duplicate_result.issues["error_code"].eq("DUPLICATE_COLUMN_NAME").any()

        strategy_data = pd.DataFrame(
            {
                "constant": [None, "x", None],
                "mean": [1.0, None, 3.0],
                "median": [1.0, None, 5.0],
                "mode": ["m", None, "m"],
                "forward": ["a", None, "b"],
                "backward": [None, "b", "c"],
                "none": [None, "x", None],
            }
        )
        strategy_data = _ensure_internal_row_numbers(strategy_data)
        strategy_config = WorkflowConfig(
            missing_value_strategies={
                "constant": {"method": "constant", "value": "filled"},
                "mean": {"method": "mean"},
                "median": {"method": "median"},
                "mode": {"method": "mode"},
                "forward": {"method": "forward_fill"},
                "backward": {"method": "backward_fill"},
                "none": {"method": "none"},
            }
        )
        strategy_filled, strategy_summary = apply_missing_value_strategies(
            strategy_data, strategy_config, ChangeTracker(100)
        )
        for column in ["constant", "mean", "median", "mode", "forward", "backward"]:
            assert strategy_filled[column].isna().sum() == 0, f"{column} 策略未正确填充"
        assert strategy_filled["none"].isna().sum() == 2
        assert set(strategy_summary["strategy"]) >= {
            "constant", "mean", "median", "mode", "forward_fill", "backward_fill", "none"
        }

        # 小规模 CSV / Excel 读取测试：文件只存在于临时目录，自检结束自动清理。
        smoke_source = source.head(3)
        csv_path = temporary_path / "smoke.csv"
        xlsx_path = temporary_path / "smoke.xlsx"
        smoke_source.to_csv(csv_path, index=False, encoding="utf-8-sig")
        smoke_source.to_excel(xlsx_path, index=False, engine="openpyxl")
        csv_read = read_input_file(csv_path, config)
        excel_read = read_input_file(xlsx_path, config)
        assert csv_read.dataframe[INTERNAL_ROW_NUMBER].tolist() == [2, 3, 4]
        assert excel_read.dataframe[INTERNAL_ROW_NUMBER].tolist() == [2, 3, 4]
        assert csv_read.used_encoding in {"utf-8-sig", "utf-8"}

        report_path = temporary_path / "self_test_report.xlsx"
        summary = build_summary_dataframe(
            result,
            input_path="self_test_memory",
            output_path=report_path,
            started_at=datetime.now(),
            elapsed_seconds=0.0,
        )
        write_excel_report(
            report_path,
            result,
            summary,
            pd.DataFrame(log_records),
            logger,
        )
        assert report_path.exists() and report_path.stat().st_size > 0
        workbook = load_workbook(report_path, read_only=True, data_only=True)
        required_sheets = {
            "Summary",
            "Clean_Data",
            "Warning_Data",
            "Rejected_Data",
            "Validation_Issues",
            "Missing_Summary",
            "Duplicate_Summary",
            "Type_Conversion",
            "Value_Distribution",
            "Outlier_Summary",
            "Column_Profile",
            "Run_Log",
        }
        assert required_sheets.issubset(set(workbook.sheetnames)), "Excel 报告缺少规定工作表"
        workbook.close()


# =============================================================================
# 15. main() 运行入口
# =============================================================================


def _parse_sheet_name(value: str) -> int | str:
    """把纯整数命令行工作表参数转换为索引，其余保留为名称。"""
    stripped = value.strip()
    if re.fullmatch(r"-?\d+", stripped):
        return int(stripped)
    return stripped


def _build_argument_parser() -> argparse.ArgumentParser:
    """创建命令行参数解析器。"""
    parser = argparse.ArgumentParser(description="日常 CSV/Excel 数据清洗与验证工作流")
    parser.add_argument("--input", help="输入 .csv/.xlsx/.xls 文件路径")
    parser.add_argument("--output", help="输出目录")
    parser.add_argument("--sheet-name", help="Excel 工作表名称或从 0 开始的索引")
    parser.add_argument("--encoding", help="CSV 编码；未指定时自动回退检测")
    parser.add_argument("--self-test", action="store_true", help="运行内置自检")
    return parser


def _timestamped_report_path(output_dir: Path) -> Path:
    """生成秒级时间戳报告名；同秒重名时添加递增后缀。"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    candidate = output_dir / f"data_quality_report_{timestamp}.xlsx"
    suffix = 2
    while candidate.exists():
        candidate = output_dir / f"data_quality_report_{timestamp}_{suffix}.xlsx"
        suffix += 1
    return candidate


def _apply_cli_overrides(config: WorkflowConfig, args: argparse.Namespace) -> WorkflowConfig:
    """使用命令行参数覆盖顶部默认配置。"""
    if args.input:
        config.input_path = Path(args.input).expanduser().resolve()
    if args.output:
        config.output_dir = Path(args.output).expanduser().resolve()
    if args.sheet_name is not None:
        config.sheet_name = _parse_sheet_name(args.sheet_name)
    if args.encoding:
        config.csv_encoding = args.encoding
    return config


def main() -> None:
    """解析配置并执行完整文件工作流；关键错误以非零状态退出。"""
    started_at = datetime.now()
    started_counter = time.perf_counter()
    logger, log_records = configure_logging(console=True)
    logger.info("程序开始")
    try:
        args = _build_argument_parser().parse_args()
        config = _apply_cli_overrides(WorkflowConfig.from_user_config(), args)

        if RUN_SELF_TEST or args.self_test:
            logger.info("内置自检开始")
            run_self_test()
            logger.info("内置自检通过")

        if config.input_path is None:
            if RUN_SELF_TEST or args.self_test:
                logger.info("未配置输入文件，本次仅运行自检。")
                logger.info("总运行时间：%.3f 秒", time.perf_counter() - started_counter)
                return
            raise ValueError(
                "未指定输入文件。请修改脚本顶部 INPUT_PATH，或使用 --input 参数。"
            )

        logger.info("输入文件：%s", config.input_path)
        read_result = read_input_file(config.input_path, config)
        original_dataframe = read_result.dataframe.copy(deep=True)
        logger.info(
            "文件读取完成：行数=%s，列数=%s，编码=%s，工作表=%s",
            len(original_dataframe),
            len(_data_columns(original_dataframe)),
            read_result.used_encoding or "不适用",
            read_result.sheet_name if read_result.sheet_name is not None else "CSV",
        )

        result = run_dataframe_workflow(
            original_dataframe,
            config,
            source_name=config.input_path.name,
            sheet_name=read_result.sheet_name,
            logger=logger,
        )

        config.output_dir.mkdir(parents=True, exist_ok=True)
        report_path = _timestamped_report_path(config.output_dir)
        elapsed_before_write = time.perf_counter() - started_counter
        summary = build_summary_dataframe(
            result,
            input_path=config.input_path,
            output_path=report_path,
            started_at=started_at,
            elapsed_seconds=elapsed_before_write,
        )
        logger.info("准备输出 Excel 报告：%s", report_path)
        write_excel_report(report_path, result, summary, pd.DataFrame(log_records), logger)
        logger.info("Excel 报告生成完成：%s", report_path)

        csv_paths: list[Path] = []
        if config.export_csv_files:
            csv_paths = export_csv_outputs(config.output_dir, result)
            logger.info("CSV 输出完成：%s", ", ".join(map(str, csv_paths)))

        total_elapsed = time.perf_counter() - started_counter
        logger.info("总运行时间：%.3f 秒", total_elapsed)
        logger.info(
            "运行摘要：原始=%s，清洗后=%s，clean=%s，warning=%s，rejected=%s，问题=%s，质量分数=%.2f",
            len(result.original_data),
            len(result.cleaned_data),
            len(result.clean_data),
            len(result.warning_data),
            len(result.rejected_data),
            len(result.issues),
            calculate_data_quality_score(result.issues, len(result.original_data)),
        )
        # 二次更新日志表，把“报告完成”和“总运行时间”也写进工作簿。
        update_run_log_sheet(report_path, log_records, logger)
    except FileNotFoundError as exc:
        logger.critical("文件不存在：%s", exc)
        raise SystemExit(2) from exc
    except PermissionError as exc:
        logger.critical("权限错误：请关闭占用中的文件并检查目录读写权限。详细信息：%s", exc)
        raise SystemExit(3) from exc
    except UnicodeDecodeError as exc:
        logger.critical(
            "CSV 编码识别失败：已尝试配置编码及 utf-8-sig、utf-8、gbk、cp932。详细信息：%s",
            exc,
        )
        raise SystemExit(4) from exc
    except ExcelWriteError as exc:
        logger.critical("Excel 写出错误：%s", exc)
        raise SystemExit(5) from exc
    except ValueError as exc:
        logger.critical("配置或数据错误：%s", exc)
        raise SystemExit(6) from exc
    except ImportError as exc:
        logger.critical("依赖或 Excel 引擎错误：%s", exc)
        raise SystemExit(7) from exc
    except OSError as exc:
        logger.critical("文件系统错误：%s", exc)
        raise SystemExit(8) from exc
    except Exception as exc:
        logger.exception("未预期的关键错误：%s", exc)
        raise SystemExit(1) from exc


if __name__ == "__main__":
    main()
